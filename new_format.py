excel_report_generator.py

Reusable Excel report generator for SAB analytics.

Usage:
    from excel_report_generator import create_excel_report

    create_excel_report(analysis_data, "report.xlsx")

`analysis_data` should be the same list of tuples currently passed to
create_csv_report().
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    'Universal_ID', 'Question', 'Sub_Category',
    'Actual_Weightage', 'Evaluation_Score',
    'Reason', 'Final_verdict',
    'Business_verdict', 'Comparison',
    'Problem_Category', 'Product_Feature', 'Summary'
]


def _style(cell, fill=None, bold=False):
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="center",
                               vertical="center",
                               wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin,
                         top=thin, bottom=thin)


def create_excel_report(analysis_data, output_file):
    df = pd.DataFrame(analysis_data, columns=COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive_Report"

    # Detailed sheet
    detail = wb.create_sheet("Detailed_Report")
    for c, col in enumerate(df.columns, 1):
        detail.cell(row=1, column=c).value = col
        _style(detail.cell(row=1, column=c), fill="D9EAD3", bold=True)

    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            detail.cell(row=r, column=c).value = val

    uid = str(df.iloc[0]["Universal_ID"])
    problem = df.iloc[0]["Problem_Category"]
    product = df.iloc[0]["Product_Feature"]
    summary = df.iloc[0]["Summary"]

    fixed = [
        ("Universal ID", uid),
        ("Problem Category", problem),
        ("Product Feature", product),
        ("Summary", summary)
    ]

    for i, (hdr, _) in enumerate(fixed, 1):
        ws.merge_cells(start_row=1, start_column=i,
                       end_row=3, end_column=i)
        c = ws.cell(row=1, column=i)
        c.value = hdr
        _style(c, fill="4F81BD", bold=True)

    row = 4
    col = len(fixed) + 1

    categories = list(dict.fromkeys(df["Sub_Category"]))

    for cat in categories:
        temp = df[df["Sub_Category"] == cat]
        start = col

        for _, q in temp.iterrows():
            ws.merge_cells(start_row=2,
                           start_column=col,
                           end_row=2,
                           end_column=col + 1)

            qc = ws.cell(row=2, column=col)
            qc.value = str(q["Question"]).replace("\n", " ").strip()
            _style(qc, fill="BDD7EE", bold=True)

            ws.cell(row=3, column=col).value = "Weightage"
            ws.cell(row=3, column=col + 1).value = "Reason"

            _style(ws.cell(row=3, column=col), fill="D9EAD3", bold=True)
            _style(ws.cell(row=3, column=col + 1), fill="D9EAD3", bold=True)

            ws.cell(row=row, column=col).value = q["Final_verdict"]
            ws.cell(row=row, column=col + 1).value = q["Reason"]

            ws.cell(row=row, column=col + 1).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

            col += 2

        ws.merge_cells(start_row=1,
                       start_column=start,
                       end_row=1,
                       end_column=col - 1)

        cc = ws.cell(row=1, column=start)
        cc.value = cat.replace("_", " ")
        _style(cc, fill="9DC3E6", bold=True)

    vals = [uid, problem, product, summary]
    for i, v in enumerate(vals, 1):
        ws.cell(row=row, column=i).value = v

    for r in ws.iter_rows():
        for cell in r:
            if cell.value is not None:
                thin = Side(style="thin")
                cell.border = Border(left=thin, right=thin,
                                     top=thin, bottom=thin)

    widths = {
        1:18,
        2:20,
        3:20,
        4:45
    }

    for c in range(5, ws.max_column + 1):
        if c % 2 == 1:
            widths[c] = 10
        else:
            widths[c] = 45

    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "E4"

    wb.save(output_file)

    return df
